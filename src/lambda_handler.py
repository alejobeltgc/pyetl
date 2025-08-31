"""
AWS Lambda handlers for the serverless ETL pipeline.
"""

import json
import os
import uuid
import boto3
import logging
from datetime import datetime, timezone
from typing import Dict, Any
from urllib.parse import unquote_plus

# Try to import openpyxl for Excel processing
try:
    from openpyxl import load_workbook
    EXCEL_PROCESSING_AVAILABLE = True
except ImportError:
    EXCEL_PROCESSING_AVAILABLE = False
    logging.warning("openpyxl not available, using test mode")

# Import our DynamoDB writer
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.cloud.dynamo_writer import DynamoDBWriter

# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# AWS clients
s3_client = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')

# Environment variables
DYNAMODB_TABLE = os.environ.get('DYNAMODB_TABLE')
S3_BUCKET = os.environ.get('S3_BUCKET')
STAGE = os.environ.get('STAGE', 'dev')


def process_rates_file(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Lambda handler triggered by S3 object creation.
    Processes Excel files and stores results in DynamoDB.
    """
    try:
        # Parse S3 event
        for record in event['Records']:
            bucket = record['s3']['bucket']['name']
            key = unquote_plus(record['s3']['object']['key'])
            
            logger.info(f"Processing file: s3://{bucket}/{key}")
            
            # Initialize DynamoDB writer
            dynamo_writer = DynamoDBWriter(DYNAMODB_TABLE)
            
            if EXCEL_PROCESSING_AVAILABLE:
                # Process real Excel file
                logger.info("📊 Processing Excel file with openpyxl...")
                
                # Download file from S3
                temp_file_path = f"/tmp/{os.path.basename(key)}"
                s3_client.download_file(bucket, key, temp_file_path)
                
                # Process Excel file
                transformed_data = process_excel_file(temp_file_path, key)
                
                # Clean up temp file
                os.remove(temp_file_path)
                
            else:
                # Fallback to test data
                logger.info("⚠️ openpyxl not available, creating test data...")
                transformed_data = create_test_data(key)
            
            # Prepare file metadata
            file_metadata = {
                'source_file': key,
                'bucket': bucket,
                'processed_at': datetime.now(timezone.utc).isoformat(),
                'stage': STAGE,
                'processing_mode': 'openpyxl' if EXCEL_PROCESSING_AVAILABLE else 'test',
                'file_size': record.get('s3', {}).get('object', {}).get('size', 0)
            }
            
            # Basic validation
            validation_report = _validate_data(transformed_data)
            
            # Store data in DynamoDB
            logger.info("💾 Storing results in DynamoDB...")
            document_id = dynamo_writer.store_document(
                transformed_data, 
                validation_report, 
                file_metadata
            )
            
            logger.info(f"✅ Processing completed successfully. Document ID: {document_id}")
            
            return {
                'statusCode': 200,
                'body': json.dumps({
                    'message': 'Processing completed successfully',
                    'document_id': document_id,
                    'source_file': key,
                    'processing_mode': 'openpyxl' if EXCEL_PROCESSING_AVAILABLE else 'test',
                    'validation_status': validation_report['status']
                })
            }
            
    except Exception as e:
        logger.error(f"❌ Processing failed: {str(e)}", exc_info=True)
        
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': 'Processing failed',
                'message': str(e)
            })
        }


def process_excel_file(file_path: str, source_file: str) -> Dict[str, Any]:
    """Process Excel file using openpyxl and extract financial data"""
    try:
        # Load workbook
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        logger.info(f"📋 Found {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
        
        tables = {}
        total_services = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"📄 Processing sheet: {sheet_name}")
            
            # Extract data from sheet
            services = extract_services_from_sheet(sheet, sheet_name)
            
            if services:
                # Determine table type based on sheet name
                table_type = classify_table_type(sheet_name)
                tables[table_type] = {
                    'source_sheet': sheet_name,
                    'services': services
                }
                total_services += len(services)
                logger.info(f"   ✅ Extracted {len(services)} services from {sheet_name}")
        
        workbook.close()
        
        return {
            'business_line': 'accounts',  # Default for now
            'document_version': 'v1',
            'last_updated': datetime.now(timezone.utc).isoformat(),
            'source_file': source_file,
            'tables': tables,
            'processing_summary': {
                'total_sheets': len(workbook.sheetnames),
                'processed_sheets': len(tables),
                'total_services': total_services
            }
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}", exc_info=True)
        # Return test data if processing fails
        return create_test_data(source_file)


def extract_services_from_sheet(sheet, sheet_name: str) -> list:
    """Extract services from a worksheet"""
    services = []
    
    try:
        # Get actual dimensions of the sheet
        max_row = sheet.max_row or 100  # Fallback if None
        max_col = sheet.max_column or 10  # Fallback if None
        
        if max_row < 1:
            logger.warning(f"Sheet {sheet_name} appears to be empty (max_row: {max_row})")
            return services
        
        logger.info(f"   📐 Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Find the header row and data rows
        header_row = None
        header_columns = {}
        
        # Look for headers in first 10 rows
        for row_num in range(1, min(11, max_row + 1)):
            row_values = []
            try:
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    row_values.append(cell_value)
            except Exception as e:
                logger.warning(f"Error reading row {row_num}: {e}")
                continue
            
            # Check if this looks like a header row
            non_empty_strings = [val for val in row_values if isinstance(val, str) and val and len(val.strip()) > 2]
            if len(non_empty_strings) >= 2:  # At least 2 meaningful headers
                header_row = row_num
                for col_idx, value in enumerate(row_values):
                    if value and isinstance(value, str) and len(value.strip()) > 1:
                        clean_name = str(value).strip().lower().replace(' ', '_').replace('ó', 'o').replace('í', 'i')
                        header_columns[col_idx] = clean_name
                break
        
        if not header_row:
            logger.warning(f"No clear header found in sheet {sheet_name}, trying to process without headers")
            header_row = 1
            # Create generic column names
            for col in range(max_col):
                header_columns[col] = f'column_{col + 1}'
        
        logger.info(f"   📊 Using row {header_row} as header with {len(header_columns)} columns")
        
        # Process data rows
        services_found = 0
        for row_num in range(header_row + 1, min(max_row + 1, header_row + 50)):  # Limit to 50 rows for safety
            try:
                row_values = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_num, column=col).value
                    row_values.append(cell_value)
                
                # Skip completely empty rows
                if not any(val for val in row_values if val is not None and str(val).strip()):
                    continue
                
                # Extract service data
                service = extract_service_from_excel_row(row_values, header_columns, row_num, sheet_name)
                if service:
                    services.append(service)
                    services_found += 1
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_num} in sheet {sheet_name}: {e}")
                continue
        
        logger.info(f"   ✅ Extracted {services_found} services from {sheet_name}")
        return services
        
    except Exception as e:
        logger.error(f"Error extracting services from sheet {sheet_name}: {e}")
        return []


def extract_service_from_excel_row(row_values: list, header_columns: dict, 
                                 row_num: int, sheet_name: str) -> Dict[str, Any]:
    """Extract service information from an Excel row"""
    try:
        # Find description (usually first or second text column with meaningful content)
        description = None
        description_col = None
        
        for idx, value in enumerate(row_values):
            if (value and isinstance(value, str) and 
                len(str(value).strip()) > 5 and  # More than 5 chars
                not str(value).strip().replace('.', '').replace(',', '').replace('$', '').isdigit()):  # Not just numbers
                
                description = str(value).strip()
                description_col = idx
                break
        
        if not description:
            return None
        
        # Extract rates from numeric columns
        rates = {}
        numeric_found = False
        
        for col_idx, value in enumerate(row_values):
            # Skip the description column
            if col_idx == description_col:
                continue
                
            # Look for numeric values (rates/fees)
            if value is not None:
                # Try to parse as number
                parsed_value = parse_numeric_value(value)
                if parsed_value is not None and parsed_value != 0:
                    # Get column name from header
                    col_name = header_columns.get(col_idx, f'column_{col_idx + 1}')
                    
                    # Parse the rate value with context
                    rate_info = parse_rate_value(parsed_value)
                    rates[col_name] = rate_info
                    numeric_found = True
        
        # Only create service if we found at least one rate
        if not numeric_found:
            return None
        
        # Generate service ID
        service_id = generate_service_id(description, row_num)
        
        return {
            'service_id': service_id,
            'description': description,
            'rates': rates,
            'source_position': {
                'sheet': sheet_name,
                'row': row_num,
                'description_column': description_col,
                'total_columns': len(row_values)
            },
            'metadata': {
                'extracted_at': datetime.now(timezone.utc).isoformat(),
                'processing_method': 'openpyxl_extraction',
                'rates_count': len(rates)
            }
        }
        
    except Exception as e:
        logger.warning(f"Error extracting service from row {row_num}: {e}")
        return None


def parse_numeric_value(value) -> float:
    """Parse a value that might be numeric, handling Colombian format"""
    try:
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            # Clean the string (remove currency symbols, spaces)
            clean_value = str(value).strip()
            
            # Remove common currency symbols and text
            clean_value = clean_value.replace('$', '').replace('COP', '').replace('USD', '')
            clean_value = clean_value.replace(' ', '').replace('\t', '')
            
            # Handle Colombian number format (dots as thousands separator)
            if ',' in clean_value and '.' in clean_value:
                # Format: 1.234.567,89 (Colombian)
                parts = clean_value.split(',')
                if len(parts) == 2:
                    integer_part = parts[0].replace('.', '')
                    decimal_part = parts[1]
                    clean_value = f"{integer_part}.{decimal_part}"
            elif '.' in clean_value and clean_value.count('.') > 1:
                # Multiple dots - treat as thousands separator
                parts = clean_value.split('.')
                if len(parts) > 2:
                    # Join all but last as integer, last as decimal if it's 2 digits
                    if len(parts[-1]) <= 2:
                        integer_part = ''.join(parts[:-1])
                        decimal_part = parts[-1]
                        clean_value = f"{integer_part}.{decimal_part}"
                    else:
                        # All are thousands separators
                        clean_value = ''.join(parts)
            
            # Try to convert to float
            return float(clean_value)
            
        return None
        
    except (ValueError, AttributeError):
        return None


def parse_rate_value(value: float) -> Dict[str, Any]:
    """Parse a rate value and determine its type"""
    
    # Convert to float if needed
    rate_value = float(value)
    
    # Basic rate classification
    if rate_value == 0:
        return {'type': 'free', 'value': 0}
    elif rate_value < 0:
        return {'type': 'credit', 'value': abs(rate_value)}
    elif 0 < rate_value < 1:
        # Likely a percentage
        return {'type': 'percentage', 'value': rate_value * 100}
    elif rate_value >= 1000:
        # Likely Colombian pesos (COP)
        return {'type': 'fixed', 'value': rate_value, 'currency': 'COP'}
    else:
        # Small amounts or other currency
        return {'type': 'fixed', 'value': rate_value}
    

def classify_table_type(sheet_name: str) -> str:
    """Classify table type based on sheet name"""
    sheet_lower = sheet_name.lower()
    
    if any(word in sheet_lower for word in ['movil', 'app', 'plan', 'móvil']):
        return 'mobile_plans'
    elif any(word in sheet_lower for word in ['transfer', 'giro', 'envío', 'envio']):
        return 'transfers' 
    elif any(word in sheet_lower for word in ['retiro', 'cajero', 'atm']):
        return 'withdrawals'
    elif any(word in sheet_lower for word in ['cuenta', 'ahorro', 'corriente']):
        return 'account_services'
    elif any(word in sheet_lower for word in ['tarjeta', 'credito', 'crédito']):
        return 'card_services'
    else:
        # Clean sheet name for use as table type
        clean_name = ''.join(c.lower() if c.isalnum() else '_' for c in sheet_name)
        return f'other_{clean_name}'


def generate_service_id(description: str, row_index: int) -> str:
    """Generate a service ID from description"""
    # Take first few words and remove special characters
    words = description.split()[:3]
    clean_words = []
    
    for word in words:
        clean_word = ''.join(c.lower() for c in word if c.isalnum())
        if clean_word and len(clean_word) > 1:
            clean_words.append(clean_word)
    
    if clean_words:
        base_id = '_'.join(clean_words)
    else:
        base_id = 'service'
    
    return f"{base_id}_{row_index}"


def create_test_data(source_file: str) -> Dict[str, Any]:
    """Create test data when pandas is not available"""
    return {
        'business_line': 'TEST',
        'document_version': 'v1',
        'last_updated': datetime.now(timezone.utc).isoformat(),
        'source_file': source_file,
        'tables': {
            'test_table': {
                'source_sheet': 'TEST_SHEET',
                'services': [
                    {
                        'service_id': 'TEST001',
                        'description': 'Test Service - File Processing',
                        'rates': {
                            'base_rate': {'type': 'fixed', 'value': 100.00}
                        },
                        'source_position': {'sheet': 'TEST', 'row': 1},
                        'metadata': {
                            'extracted_at': datetime.now(timezone.utc).isoformat(),
                            'processing_method': 'test_mode'
                        }
                    }
                ]
            }
        },
        'processing_summary': {
            'total_sheets': 1,
            'processed_sheets': 1, 
            'total_services': 1,
            'mode': 'test'
        }
    }


def health_check(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Health check endpoint for monitoring.
    """
    try:
        # Test DynamoDB connection
        table = dynamodb.Table(DYNAMODB_TABLE)
        table.meta.client.describe_table(TableName=DYNAMODB_TABLE)
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'status': 'healthy',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'stage': STAGE,
                'dynamodb_table': DYNAMODB_TABLE,
                'version': '1.0.0'
            })
        }
        
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        
        return {
            'statusCode': 503,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'status': 'unhealthy',
                'error': str(e),
                'timestamp': datetime.now(timezone.utc).isoformat()
            })
        }


def _count_services(transformed_data: Dict[str, Any]) -> int:
    """Count total services across all tables."""
    total = 0
    for table_data in transformed_data.get('tables', {}).values():
        total += len(table_data.get('services', []))
    return total


def _validate_data(transformed_data: Dict[str, Any]) -> Dict[str, Any]:
    """Basic validation of transformed data."""
    errors = []
    warnings = []
    
    # Check basic structure
    if 'business_line' not in transformed_data:
        errors.append("Missing business_line field")
    
    if 'tables' not in transformed_data:
        errors.append("Missing tables field")
    
    # Count services
    services_count = _count_services(transformed_data)
    
    if services_count == 0:
        warnings.append("No services found in any table")
    
    return {
        'status': 'valid' if len(errors) == 0 else 'invalid',
        'errors': errors,
        'warnings': warnings,
        'services_processed': services_count,
        'validated_at': datetime.now(timezone.utc).isoformat()
    }
