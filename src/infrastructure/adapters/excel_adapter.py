"""Excel processing adapter using openpyxl."""

import logging
from typing import BinaryIO, List, Dict, Any, Optional
from decimal import Decimal
import re

try:
    from openpyxl import load_workbook
    EXCEL_PROCESSING_AVAILABLE = True
except ImportError:
    EXCEL_PROCESSING_AVAILABLE = False

from ...domain.entities import Document, FinancialService
from ...domain.services import ExcelProcessorService

logger = logging.getLogger(__name__)


class OpenpyxlExcelAdapter:
    """
    Openpyxl implementation for Excel processing.
    
    Concrete implementation that handles Excel parsing using openpyxl library.
    """
    
    def __init__(self):
        if not EXCEL_PROCESSING_AVAILABLE:
            raise ExcelAdapterError("openpyxl is not available")
        
        self.excel_processor = ExcelProcessorService()
    
    async def process_excel_file(self, file_content: BinaryIO, filename: str, 
                                document_id: str) -> Document:
        """
        Process Excel file content and extract financial services.
        
        Args:
            file_content: Excel file content
            filename: Original filename
            document_id: Unique document identifier
            
        Returns:
            Document with extracted services
        """
        try:
            # Load workbook
            workbook = load_workbook(file_content, data_only=True)
            
            # Determine business line from sheets
            business_line = self._determine_business_line(workbook.sheetnames)
            
            # Create document
            document = Document(
                document_id=document_id,
                business_line=business_line,
                filename=filename
            )
            
            # Add processing metadata
            document.processing_metadata = {
                'sheets_found': workbook.sheetnames,
                'processing_method': 'openpyxl',
                'excel_format': 'xlsx'
            }
            
            # Process each relevant sheet
            relevant_sheets = self._get_relevant_sheets(workbook.sheetnames)
            logger.info(f"📋 Found {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
            logger.info(f"🎯 Processing {len(relevant_sheets)} relevant sheets: {relevant_sheets}")
            
            for sheet_name in relevant_sheets:
                try:
                    sheet = workbook[sheet_name]
                    services = await self._extract_services_from_sheet(
                        sheet, sheet_name, document_id
                    )
                    
                    for service in services:
                        document.add_service(service)
                    
                    logger.info(f"✅ Extracted {len(services)} services from {sheet_name}")
                    
                except Exception as e:
                    logger.error(f"❌ Error processing sheet {sheet_name}: {str(e)}")
                    document.processing_metadata[f'error_{sheet_name}'] = str(e)
            
            logger.info(f"📊 Total services extracted: {document.get_service_count()}")
            
            return document
            
        except Exception as e:
            raise ExcelAdapterError(f"Failed to process Excel file {filename}: {str(e)}") from e
    
    def _determine_business_line(self, sheet_names: List[str]) -> str:
        """Determine the primary business line from sheet names."""
        # Count business line indicators
        business_line_counts = {
            'accounts': 0,
            'loans': 0,
            'other': 0
        }
        
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            if any(term in sheet_lower for term in ['tarifa', 'limite', 'cuenta']):
                business_line_counts['accounts'] += 1
            elif any(term in sheet_lower for term in ['tasa', 'credito', 'prestamo']):
                business_line_counts['loans'] += 1
            else:
                business_line_counts['other'] += 1
        
        # Return the most common business line
        return max(business_line_counts, key=business_line_counts.get)
    
    def _get_relevant_sheets(self, sheet_names: List[str]) -> List[str]:
        """Filter to only relevant sheets for processing."""
        relevant_patterns = [
            r'tarifa', r'limite', r'tasa', r'rate', r'fee', r'cost'
        ]
        
        relevant_sheets = []
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()
            if any(re.search(pattern, sheet_lower) for pattern in relevant_patterns):
                relevant_sheets.append(sheet_name)
        
        return relevant_sheets
    
    async def _extract_services_from_sheet(self, sheet, sheet_name: str, 
                                          document_id: str) -> List[FinancialService]:
        """Extract financial services from a single sheet."""
        services = []
        
        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row is None or max_row == 0:
            logger.warning(f"Sheet {sheet_name} appears to be empty")
            return services
        
        logger.info(f"Processing sheet {sheet_name}: {max_row} rows × {max_col} columns")
        
        # Find header row
        header_row_idx = self._find_header_row(sheet, max_row)
        if header_row_idx is None:
            logger.warning(f"No header row found in sheet {sheet_name}")
            return services
        
        # Get headers
        headers = self._extract_headers(sheet, header_row_idx, max_col)
        logger.info(f"Headers found: {headers}")
        
        # Extract data rows
        for row_idx in range(header_row_idx + 1, max_row + 1):
            try:
                service = self._extract_service_from_row(
                    sheet, row_idx, headers, sheet_name, document_id, row_idx - header_row_idx
                )
                
                if service and service.description.strip():
                    services.append(service)
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_idx} in {sheet_name}: {str(e)}")
        
        return services
    
    def _find_header_row(self, sheet, max_row: int) -> Optional[int]:
        """Find the row that contains column headers."""
        header_indicators = [
            'servicio', 'descripcion', 'concepto', 'plan', 'tarifa', 'limite', 'tasa',
            'service', 'description', 'concept', 'rate', 'fee', 'limit'
        ]
        
        for row_idx in range(1, min(10, max_row + 1)):  # Check first 10 rows
            row_values = []
            for col_idx in range(1, min(10, sheet.max_column + 1)):  # Check first 10 columns
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_values.append(str(cell_value).lower().strip())
            
            # Check if this row contains header indicators
            header_count = sum(1 for header in header_indicators 
                              if any(header in val for val in row_values))
            
            if header_count >= 1:  # At least one header indicator
                return row_idx
        
        return 1  # Default to first row
    
    def _extract_headers(self, sheet, header_row_idx: int, max_col: int) -> List[str]:
        """Extract column headers from the header row."""
        headers = []
        for col_idx in range(1, max_col + 1):
            cell_value = sheet.cell(row=header_row_idx, column=col_idx).value
            header = str(cell_value).strip() if cell_value else f"column_{col_idx}"
            headers.append(header)
        
        return headers
    
    def _extract_service_from_row(self, sheet, row_idx: int, headers: List[str], 
                                 sheet_name: str, document_id: str, 
                                 service_index: int) -> Optional[FinancialService]:
        """Extract a financial service from a single row."""
        # Get row data
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            row_data[header] = cell_value
        
        # Find description field
        description = self._find_description(row_data)
        if not description or description.strip() == '':
            return None
        
        # Create service
        table_type = f"other_{sheet_name.lower()}"
        business_line = self.excel_processor.classify_business_line(table_type)
        service_id = self.excel_processor.generate_service_id(
            description, table_type, service_index
        )
        
        service = FinancialService(
            service_id=service_id,
            description=description,
            business_line=business_line,
            table_type=table_type,
            document_id=document_id,
            source_position={
                'sheet': sheet_name,
                'row': row_idx,
                'header_row': headers
            }
        )
        
        # Extract rates from other columns
        for header, value in row_data.items():
            if header != description and value is not None:
                # Skip empty values and description field
                if str(value).strip() == '' or header in ['descripcion', 'servicio', 'concepto']:
                    continue
                
                try:
                    rate = self.excel_processor.create_rate_from_value(value, header)
                    plan_name = self._normalize_plan_name(header)
                    service.add_rate(plan_name, rate)
                except Exception as e:
                    logger.warning(f"Error creating rate for {header}: {str(e)}")
        
        return service
    
    def _find_description(self, row_data: Dict[str, Any]) -> Optional[str]:
        """Find the description field in the row data."""
        description_patterns = [
            'descripcion', 'servicio', 'concepto', 'description', 'service', 'concept'
        ]
        
        # First, try exact matches
        for pattern in description_patterns:
            for header, value in row_data.items():
                if pattern.lower() in header.lower() and value:
                    return str(value).strip()
        
        # Then try first non-empty value
        for header, value in row_data.items():
            if value and str(value).strip():
                return str(value).strip()
        
        return None
    
    def _normalize_plan_name(self, header: str) -> str:
        """Normalize plan/column name."""
        # Clean and normalize header name
        normalized = re.sub(r'[^\w\s]', ' ', header.lower())
        normalized = re.sub(r'\s+', '_', normalized.strip())
        return normalized


class ExcelAdapterError(Exception):
    """Raised when Excel adapter operations fail."""
    pass
