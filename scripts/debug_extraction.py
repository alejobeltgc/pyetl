#!/usr/bin/env python3
"""
Script de debugging para analizar por qué se extraen pocos datos
"""
import sys
import asyncio
from pathlib import Path
from openpyxl import load_workbook

# Agregar path para imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.domain.strategies.strategy_factory import ExtractionStrategyFactory

def analyze_excel_file(file_path: str):
    """Analiza archivo Excel para debugging"""
    print(f"🔍 ANÁLISIS DETALLADO: {Path(file_path).name}")
    print("=" * 60)
    
    # Cargar archivo
    workbook = load_workbook(file_path, data_only=True)
    
    print(f"📊 Sheets encontrados: {workbook.sheetnames}")
    
    # Analizar cada sheet
    for sheet_name in workbook.sheetnames:
        print(f"\n📋 SHEET: {sheet_name}")
        print("-" * 40)
        
        sheet = workbook[sheet_name]
        
        # Información básica
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        print(f"   📏 Dimensiones: {max_row} filas x {max_col} columnas")
        
        # Convertir a array
        sheet_data = []
        for row_idx in range(1, min(max_row + 1, 21)):  # Solo primeras 20 filas
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            sheet_data.append(row_data)
        
        # Mostrar primeras filas
        print(f"   📄 Primeras filas con datos:")
        for i, row in enumerate(sheet_data[:10]):
            # Filtrar celdas vacías para mostrar
            non_empty = [str(cell)[:50] for cell in row if cell is not None and str(cell).strip()]
            if non_empty:
                print(f"      Fila {i+1}: {non_empty[:5]}{'...' if len(non_empty) > 5 else ''}")
        
        # Analizar con strategy
        factory = ExtractionStrategyFactory()
        business_line = factory.detect_business_line(workbook.sheetnames, Path(file_path).name)
        strategy = factory.get_strategy(business_line)
        
        print(f"   🎯 Business line detectado: {business_line}")
        print(f"   🔧 Strategy: {strategy.__class__.__name__}")
        print(f"   ✅ Should process: {strategy.should_process_sheet(sheet_name)}")
        print(f"   📂 Sheet type: {strategy.classify_sheet_type(sheet_name)}")
        
        if strategy.should_process_sheet(sheet_name):
            # Análisis de headers
            data_start_row = strategy.find_data_start_row(sheet_data, sheet_name)
            print(f"   🎯 Data start row: {data_start_row}")
            
            if data_start_row:
                headers = strategy.extract_headers(sheet_data, data_start_row)
                print(f"   📋 Headers ({len(headers)}): {headers[:5]}{'...' if len(headers) > 5 else ''}")
                
                # Contar filas con datos real
                data_rows = 0
                valid_services = 0
                
                for row_idx in range(data_start_row, len(sheet_data)):
                    row_data = sheet_data[row_idx]
                    
                    # Verificar si la fila tiene datos
                    has_data = any(cell is not None and str(cell).strip() for cell in row_data)
                    if has_data:
                        data_rows += 1
                        
                        # Probar extraer servicio
                        try:
                            service = strategy.extract_service_from_row(
                                row_data, headers, sheet_name, row_idx, "debug-test"
                            )
                            if service:
                                valid_services += 1
                            elif data_rows <= 5:  # Mostrar solo primeras 5 para debug
                                print(f"      ❌ Fila {row_idx+1} no generó servicio: {[str(c)[:30] for c in row_data[:3]]}")
                        except Exception as e:
                            if data_rows <= 5:
                                print(f"      💥 Error en fila {row_idx+1}: {e}")
                
                print(f"   📊 Filas con datos: {data_rows}")
                print(f"   ✅ Servicios válidos: {valid_services}")
                print(f"   📈 Tasa de éxito: {(valid_services/data_rows*100):.1f}%" if data_rows > 0 else "   📈 Tasa de éxito: 0%")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python debug_extraction.py <archivo_excel>")
        sys.exit(1)
    
    analyze_excel_file(sys.argv[1])
